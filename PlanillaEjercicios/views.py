
# from django.http import HttpResponse
from msilib.schema import ListView
from pyexpat import model
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.db.models import Q
from .models import Ejercicios, HabilidadEjercicio, Habilidades
from django.views.generic import ListView
import openpyxl
from django.contrib import messages

class EjercicioListView(ListView):
	model = Ejercicios
	template_name = 'gestionEjercicios.html'
	
	def get_context_data(self, **kwargs):
		context = super().get_context_data(**kwargs)
		context['titulo'] = 'Gestion de Ejercicios'
		return context

def registro_ejercicio(request):
	data = {
		'titulo':'Edicion de Ejercicios',

	}
	return render(request,"registroEjercicio.html")

def creacion_ejercicio(request):
	excel_file 	= request.FILES['file']
	wb = openpyxl.load_workbook(filename = excel_file)
	sheets = wb.sheetnames
	worksheet = wb[sheets[0]]
	fila_habilidades = list()
	dic = {'F':'Abstraction','G':'Abstraction','H':'Abstraction','I':'Decomposition','J':'Decomposition','K':'Decomposition','L':'Decomposition','M':'Pattern Recognition and Generalization','N':'Pattern Recognition and Generalization','O':'Pattern Recognition and Generalization','P':'Pattern Recognition and Generalization','Q':'Algorithmic thinking','R':'Algorithmic thinking','S':'Algorithmic thinking','T':'Algorithmic thinking','U':'Algorithmic thinking'}

	#recorrer columna 2
	for row in worksheet.iter_rows(min_row=2,min_col=6,max_row=2):
		for cell in row:
			fila_habilidades.append(str(cell.value))
			try:
				existe_habilidad = Habilidades.objects.get(hablidad=cell.value)
			except:
				existe_habilidad = None

			if not existe_habilidad:
				hablidad = Habilidades(hablidad=cell.value,description=dic[cell.column_letter])
				hablidad.save()

	#recorrer desde la columna 3
	for row in worksheet.iter_rows(min_row=3,min_col=3):
		datos = list()
		for cell in row:
			datos.append(cell.value)

		#obtener Ejercicio
		try:
			ejercicio = Ejercicios.objects.get(titulo=datos[0])
		except:
			ejercicio = None

		if ejercicio:
			del datos[0:3]
			for i in range(len(fila_habilidades)):
				#obtener habilidad
				habilidad = Habilidades.objects.get(hablidad=fila_habilidades[i])

				#guardar en tabla HabilidadEjercicio
				habilidadejercicio = HabilidadEjercicio(ejercicio=ejercicio,habilidades=habilidad,valor=datos[i])
				habilidadejercicio.save()
		else:
			ejercicio = Ejercicios(titulo=datos[0],descripcion=datos[1],dificultad=datos[2])
			ejercicio.save()
			
			del datos[0:3]
			for i in range(len(fila_habilidades)):
				#obtener habilidad
				habilidad = Habilidades.objects.get(hablidad=fila_habilidades[i])

				#guardar en tabla HabilidadEjercicio
				habilidadejercicio = HabilidadEjercicio(ejercicio=ejercicio,habilidades=habilidad,valor=datos[i])
				habilidadejercicio.save()

	ejercicios = Ejercicios.objects.all()
	messages.success(request, 'Planilla carga de forma exitosa!.')
	return render(request,'ejercicios.html',{'ejercicios':ejercicios})

def ejercicio(request):
	#obtener todos los ejercicios
	ejercicios = Ejercicios.objects.all()

	return render(request,'ejercicios.html',{'ejercicios':ejercicios})

def editar_ejercicio(request,ejercicio_id):
	if request.method == 'POST':
		#sactualizar atributos de Ejercicio
		Ejercicios.objects.filter(pk=ejercicio_id).update(titulo=request.POST['txtTitulo'],descripcion=request.POST['txtDescripcion'],dificultad=request.POST['numDificultad'])

		messages.success(request, 'Ejericicio modificado de forma exitosa!.')
		return redirect("ejercicio")
	else:
		ejercicio = Ejercicios.objects.get(pk=ejercicio_id)
		return render (request,'edicionEjercicios.html',{'ejercicio':ejercicio})

def ver_ejercicio(request,ejercicio_id):
	#obtener ejercicio por id
	ejercicio = Ejercicios.objects.get(pk=ejercicio_id)
	
	#declarar arreglos
	list_abstraction = []
	list_decomposition = []
	list_pattern = []
	list_algorithmic = []

	#obtener habilidades por su descripcion
	abstraction = Habilidades.objects.filter(description="Abstraction")
	decomposition = Habilidades.objects.filter(description="Decomposition")
	pattern = Habilidades.objects.filter(description="Pattern Recognition and Generalization")
	algorithmic = Habilidades.objects.filter(description="Algorithmic thinking")

	#recorrer habilidad abstraction
	for a in abstraction:
		#se define arreglo
		valores = []
		#obtiener todos los valores por habilidad y ejercicio
		habilidades_abstraction = HabilidadEjercicio.objects.filter(Q(habilidades=a) & Q(ejercicio=ejercicio))
		#recorrer HabilidadEjercicio, obtener array solo con valores y guardar un array de array en list_abstraction
		for h in habilidades_abstraction:
			valores.append(h.valor)
		list_abstraction.append(valores)

	for d in decomposition:
		valores = []
		habilidades_abstraction = HabilidadEjercicio.objects.filter(Q(habilidades=d) & Q(ejercicio=ejercicio))
		for h in habilidades_abstraction:
			valores.append(h.valor)
		list_decomposition.append(valores)
	
	for p in pattern:
		valores = []
		habilidades_abstraction = HabilidadEjercicio.objects.filter(Q(habilidades=p) & Q(ejercicio=ejercicio))
		for h in habilidades_abstraction:
			valores.append(h.valor)
		list_pattern.append(valores)
	
	for al in algorithmic:
		valores = []
		habilidades_abstraction = HabilidadEjercicio.objects.filter(Q(habilidades=al) & Q(ejercicio=ejercicio))
		for h in habilidades_abstraction:
			valores.append(h.valor)
		list_algorithmic.append(valores)

	#transponer matrices para desplegar datos ordenados
	result_abstraction = transpose(list_abstraction,abstraction.count())
	result_decomposition = transpose(list_decomposition,decomposition.count())
	result_pattern = transpose(list_pattern,pattern.count())
	result_algorithmic = transpose(list_algorithmic,algorithmic.count())

	return render(request,'verEjercicios.html',{'ejercicio':ejercicio,'result_abstraction':result_abstraction,'result_decomposition':result_decomposition,'result_pattern':result_pattern,'result_algorithmic':result_algorithmic})

def transpose(matrix,elemento):
	#operaciones matematicas para trasponer arreglos
	result = [[None for i in range(len(matrix))] for j in range(len(matrix[0]))]
	list_cabecera = []
	dict_result = []
	for i in range(len(matrix[0])):
		for j in range(len(matrix)):
			result[i][j] = matrix[j][i]

	for n in range(elemento):
		list_cabecera.append('resp'+str(n))

	#convertir arreglos a diccionario para el despliegue de datos
	for r in result:
		list_to_dict = dict(zip(list_cabecera, r))
		dict_result.append(list_to_dict)

	return dict_result

def eliminar_ejercicio(request,ejercicio_id):
	print(ejercicio_id)
	ejercicio = Ejercicios.objects.get(id=ejercicio_id)
	ejercicio.delete()
	return JsonResponse({'status': True, 'message': 'Ejecicio eliminado de forma exitosa!.'})